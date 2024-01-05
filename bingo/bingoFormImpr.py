import openpyxl

def leer_tarjetas(nombre_archivo):
    libro = openpyxl.load_workbook(nombre_archivo)

    # Crear una matriz para almacenar las combinaciones
    combinaciones = []

    # Leer la fila 1 de cada pestaña y agregarla a la matriz
    for letra in ['B', 'I', 'N', 'G', 'O']:
        hoja = libro[letra]
        combinacion = [hoja.cell(row=1, column=col).value for col in range(2, 7)]
        combinaciones.append(combinacion)

    return combinaciones

def escribir_resultado(matriz_resultado, nombre_archivo):
    libro = openpyxl.Workbook()

    # Crear pestañas para cada letra
    letras = ['B', 'I', 'N', 'G', 'O']
    for i, letra in enumerate(letras):
        hoja = libro.create_sheet(title=letra)

        # Escribir las combinaciones en la hoja de Excel
        for j, combinacion in enumerate(matriz_resultado, start=1):
            hoja.cell(row=j, column=1, value=f'Combinación {j}')
            for k, valor in enumerate(combinacion):
                hoja.cell(row=j, column=k + 2, value=valor)

    # Eliminar la hoja de inicio
    default_sheet = libro['Sheet']
    del libro['Sheet']

    libro.save(nombre_archivo)

if __name__ == "__main__":
    archivo_entrada = 'tarjetas_bingo500.xlsx'
    archivo_salida = 'resultados_combinaciones.xlsx'

    # Leer las tarjetas
    tarjetas = leer_tarjetas(archivo_entrada)

    # Transponer la matriz para obtener las combinaciones
    combinaciones = list(map(list, zip(*tarjetas)))

    # Escribir el resultado en un nuevo archivo Excel
    escribir_resultado(combinaciones, archivo_salida)
